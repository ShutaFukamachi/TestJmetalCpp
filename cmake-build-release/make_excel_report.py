# make_excel_report_nondominated_with_cmetric.py
import glob
import os
from openpyxl import Workbook

MODES = ["NOAUG_NOLS", "AUG_NOLS", "NOAUG_LS", "AUG_LS"]
POP_SIZE = 100  # 表の表示行数（100行に揃える）

def read_fun(path):
    pts = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            a, b = line.split()
            pts.append((float(a), float(b)))
    return pts

def dominates(a, b):
    # minimization, 2 objectives
    return (a[0] <= b[0] and a[1] <= b[1]) and (a[0] < b[0] or a[1] < b[1])

def extract_nondominated(points):
    nondom = []
    for i, p in enumerate(points):
        dominated = False
        for j, q in enumerate(points):
            if i != j and dominates(q, p):
                dominated = True
                break
        if not dominated:
            nondom.append(p)
    return nondom

def c_metric(A, B):
    # C(A,B) = Bの各点がAのどれかに支配される割合（%）
    if not B:
        return 0.0
    dominated_count = 0
    for b in B:
        if any(dominates(a, b) for a in A):
            dominated_count += 1
    return 100.0 * dominated_count / len(B)

def instance_base_from_fun(fun_name):
    # FUN_<base>_<mode> から base を取り出す
    s = fun_name[len("FUN_"):]
    for m in MODES:
        if s.endswith("_" + m):
            return s[:-(len(m) + 1)]
    return None

def main():
    wb = Workbook()
    wb.remove(wb.active)

    # Summary にも C-metric を入れる（任意だが便利）
    summary = wb.create_sheet("Summary")
    summary.append(["instance"] + [f"|F1({m})|" for m in MODES] + ["C(AUG_LS, NOAUG_NOLS) [%]"])

    # インスタンス列挙（NOAUG_NOLS を起点）
    fun_base = sorted(glob.glob("FUN_*_NOAUG_NOLS"))

    for f0 in fun_base:
        base = instance_base_from_fun(os.path.basename(f0))
        if base is None:
            continue

        # ---- C-metric（ログ一致用：元のFUN(100個)で計算）----
        pathA = f"FUN_{base}_AUG_LS"
        pathB = f"FUN_{base}_NOAUG_NOLS"
        A_full = read_fun(pathA) if os.path.exists(pathA) else []
        B_full = read_fun(pathB) if os.path.exists(pathB) else []
        c_val = c_metric(A_full, B_full)

        # ---- 非支配表（F1）用のデータ ----
        nd_sets = {}
        for m in MODES:
            path = f"FUN_{base}_{m}"
            if os.path.exists(path):
                pts = read_fun(path)
                nd = extract_nondominated(pts)
            else:
                nd = []
            nd_sets[m] = nd

        ws = wb.create_sheet(base[:31])

        # ヘッダ（2列×4モード=8列）
        col = 1
        for m in MODES:
            ws.cell(1, col).value = f"{m} F1 (Cmax)"
            ws.cell(1, col + 1).value = f"{m} F1 (Cost)"
            col += 2

        # 右端に Cメトリクス列を追加（これが「一番右の列」）
        last_col = 2 * len(MODES)  # 8
        ws.cell(1, last_col + 1).value = "C(AUG_LS, NOAUG_NOLS) [%]"
        ws.cell(2, last_col + 1).value = round(c_val, 2)

        # 表の中身：上からF1、足りない分は先頭F1を繰り返して100行にする
        for i in range(POP_SIZE):
            r = i + 2
            col = 1
            for m in MODES:
                nd = nd_sets[m]
                if len(nd) == 0:
                    col += 2
                    continue
                p = nd[i] if i < len(nd) else nd[0]
                ws.cell(r, col).value = p[0]
                ws.cell(r, col + 1).value = p[1]
                col += 2

        # Summary 行
        summary.append([base] + [len(nd_sets[m]) for m in MODES] + [round(c_val, 2)])

    wb.save("NSGA2_report_nondominated.xlsx")
    print("[DONE] NSGA2_report_nondominated.xlsx created")

if __name__ == "__main__":
    main()
