print("=== RUNNING NEW SCRIPT VERSION: 2026-xx-xx ===")

import glob
import os
from openpyxl import Workbook

MODES = ["NOAUG_NOLS", "AUG_NOLS", "NOAUG_LS", "AUG_LS"]
POP_SIZE = 100  # 各シートに表示する最大行数


def read_fun(path):
    pts = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            a, b = line.split()
            pts.append((float(a), float(b)))
    return pts


def dominates(a, b):
    # minimization (2 objectives)
    return (a[0] <= b[0] and a[1] <= b[1]) and (a[0] < b[0] or a[1] < b[1])


def extract_nondominated(points):
    nondom = []
    for i, p in enumerate(points):
        dom_flag = False
        for j, q in enumerate(points):
            if i != j and dominates(q, p):
                dom_flag = True
                break
        if not dom_flag:
            nondom.append(p)
    return nondom


def c_metric(A, B):
    # C(A,B) = Bの各点がAのどれかに支配される割合（%）
    if not B:
        return 0.0
    dominated_count = 0
    for b in B:
        if any(dominates(a, b) for a in A):
            dominated_count += 1
    return 100.0 * dominated_count / len(B)


def hypervolume_2d(points, ref):
    """
    2目的(min)のHVを厳密計算
    points: 非支配集合推奨
    ref: 参照点（全点より“悪い”点）
    """
    if not points:
        return 0.0

    # refより悪い点はクリップ（念のため）
    pts = [(min(p[0], ref[0]), min(p[1], ref[1])) for p in points]

    # 非支配化（念のため）
    pts = extract_nondominated(pts)

    # f1昇順に並べて右から面積を足す
    pts_sorted = sorted(pts, key=lambda x: x[0])

    hv = 0.0
    prev_x = ref[0]
    best_y = ref[1]

    for (x, y) in reversed(pts_sorted):
        if y < best_y:
            hv += (prev_x - x) * (best_y - y)
            best_y = y
            prev_x = x
        else:
            prev_x = min(prev_x, x)

    return hv


def dominated_intersection_front(A, B):
    """
    Dom(A) ∩ Dom(B) の境界を作るために
    C = { max(a,b) | a in A, b in B } を作って非支配化して返す
    """
    if not A or not B:
        return []
    C = []
    for a in A:
        for b in B:
            C.append((max(a[0], b[0]), max(a[1], b[1])))
    return extract_nondominated(C)


def hvr_intersection_over_reference(A, B, ref):
    """
    HVR = HV(Dom(A)∩Dom(B)) / HV(Dom(B)) を(%)で返す
    ここでは A=AUG_LS, B=NOAUG_NOLS を想定
    """
    B_nd = extract_nondominated(B)
    if not B_nd:
        return 0.0

    A_nd = extract_nondominated(A)
    C_nd = dominated_intersection_front(A_nd, B_nd)

    hv_B = hypervolume_2d(B_nd, ref)
    if hv_B <= 0:
        return 0.0

    hv_C = hypervolume_2d(C_nd, ref)
    return 100.0 * hv_C / hv_B


def instance_base_from_fun(fun_name):
    # FUN_<base>_<mode> から base を取り出す
    s = fun_name[len("FUN_"):]
    for m in MODES:
        if s.endswith("_" + m):
            return s[:-(len(m) + 1)]
    return None


def main():
    import os
    print("[DEBUG] script cwd =", os.getcwd())
    print("[DEBUG] script file =", __file__)


    wb = Workbook()
    wb.remove(wb.active)

    # ===== Summary =====
    summary = wb.create_sheet("Summary")
    summary.append(
        ["instance"]
        + [f"|F1({m})|" for m in MODES]
        + ["C(AUG_LS, NOAUG_NOLS) [%]"]
        + ["HVR(AUG_LS over NOAUG_NOLS) [%]"]
    )

    # NOAUG_NOLS を起点にインスタンス列挙
    fun_base = sorted(glob.glob("FUN_*_NOAUG_NOLS"))

    for f0 in fun_base:
        base = instance_base_from_fun(os.path.basename(f0))
        if base is None:
            continue

        # ---- フル集合を読み込み（C/HVR計算はフルでOK）----
        pathA = f"FUN_{base}_AUG_LS"
        pathB = f"FUN_{base}_NOAUG_NOLS"
        A_full = read_fun(pathA) if os.path.exists(pathA) else []
        B_full = read_fun(pathB) if os.path.exists(pathB) else []

        c_val = c_metric(A_full, B_full)

        # 参照点 ref（AとBの最大より“悪い”点）
        all_pts = A_full + B_full
        if all_pts:
            max_f1 = max(p[0] for p in all_pts)
            max_f2 = max(p[1] for p in all_pts)
            ref = (max_f1 * 1.05 + 1.0, max_f2 * 1.05 + 1.0)
        else:
            ref = (1.0, 1.0)

        hvr_val = hvr_intersection_over_reference(A_full, B_full, ref)

        # ---- 各モードの非支配(F1) ----
        nd_sets = {}
        for m in MODES:
            path = f"FUN_{base}_{m}"
            if os.path.exists(path):
                pts = read_fun(path)
                nd = extract_nondominated(pts)
            else:
                nd = []
            nd_sets[m] = nd

        # ===== インスタンス別シート =====
        ws = wb.create_sheet(base[:31])

        col = 1
        for m in MODES:
            ws.cell(1, col).value = f"{m} F1 (Cmax)"
            ws.cell(1, col + 1).value = f"{m} F1 (Cost)"
            col += 2

        last_col = 2 * len(MODES)  # 8
        ws.cell(1, last_col + 1).value = "C(AUG_LS, NOAUG_NOLS) [%]"
        ws.cell(2, last_col + 1).value = round(c_val, 2)
        ws.cell(1, last_col + 2).value = "HVR(AUG_LS over NOAUG_NOLS) [%]"
        ws.cell(2, last_col + 2).value = round(hvr_val, 2)

        for i in range(POP_SIZE):
            r = i + 2
            col = 1
            for m in MODES:
                nd = nd_sets[m]
                if not nd:
                    col += 2
                    continue
                p = nd[i] if i < len(nd) else nd[0]
                ws.cell(r, col).value = p[0]
                ws.cell(r, col + 1).value = p[1]
                col += 2

        # ===== Summary 行（ここが大事：HVRも入れる）=====
        summary.append(
            [base]
            + [len(nd_sets[m]) for m in MODES]
            + [round(c_val, 2)]
            + [round(hvr_val, 2)]
        )

    wb.save("NSGA2_report_nondominated.xlsx")
    print("[DONE] NSGA2_report_nondominated.xlsx created (with C-metric + HVR)")


if __name__ == "__main__":
    main()
